## image_to_excel.py
import argparse
import os

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


def main(excel_file_name, sheet_name):
    # Check if the input file exists in the current directory
    if not os.path.isfile(excel_file_name):
        print(f"エラー： '{excel_file_name}' は存在しません。")
        exit()

    # Load the selected excel file
    workbook = openpyxl.load_workbook(excel_file_name)

    # Add a new sheet with the given name
    sheet = workbook.create_sheet(sheet_name)

    # Get all directories from the current directory
    directories = sorted([d for d in os.listdir() if os.path.isdir(d)])
    # Initialize the row in which the image will be pasted
    row = 2

    # Iterate through the directories and process the images
    for directory in directories:
        # Write the directory name in the first column
        sheet.cell(column=1, row=row).value = directory
        row += 3
        # Iterate through the image files in the directory
        for image_file in sorted(os.listdir(directory)):
            if image_file.lower().endswith(".png") or image_file.lower().endswith(
                ".jpg"
            ):
                img = Image(os.path.join(directory, image_file))

                # Get the original width and height of the image
                original_width = img.width
                original_height = img.height

                # Set the maximum width and height for resizing
                max_width = 800
                max_height = 800

                # Calculate the aspect ratio of the image
                aspect_ratio = original_width / original_height

                # Determine the final width and height while preserving the aspect ratio
                if aspect_ratio > 1:
                    final_width = max_width
                    final_height = max_width / aspect_ratio
                else:
                    final_width = max_height * aspect_ratio
                    final_height = max_height

                # Set the size of the image
                img.width = final_width
                img.height = final_height

                # Paste the image in the Excel sheet
                sheet.add_image(img, f"{get_column_letter(2)}{row}")

                # Increment the row index, leaving 3 rows gap between images
                row += int(final_height) // 20 + 3

    # Save the workbook
    workbook.save(excel_file_name)

    print("Insert Images to Excel")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("excel_file_name", help="使用する既存のExcelファイル名")
    parser.add_argument("sheet_name", help="作成するシート名")
    args = parser.parse_args()
    main(args.excel_file_name, args.sheet_name)
